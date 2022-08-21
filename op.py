import openpyxl as op


save_path=r'C:\Users\xkr04\Downloads\excel\excel2.xlsx'
wb0=op.Workbook()
ws0=wb0.active
wb0.save(save_path)

wb=op.load_workbook(save_path)
ws=wb.active

lists1=['날짜','제품명','가격','수량','합계']
lists2=['2020.1.30','삼성노트북',1000000,5,'=c2*d2']
for i in range(5):
    ws.cell(row=1,column=i+1).value=lists1[i]
    ws.cell(row=2,column=i+1).value=lists2[i]
# print(wb.sheetnames)
ws.append(['2021.1.25','냉장고',8000000,6,'=c3*d3'])
ws.append(['2021.2.24','라디오',50000,8,'=c4*d4'])
wb.save(save_path)