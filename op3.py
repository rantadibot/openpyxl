import openpyxl as op
import fill_color,font_color
from openpyxl.styles import Font

wb=op.Workbook()
wb.create_sheet('Test')
ws=wb['Test']

ws.append(['순위','이름','순위','이름','순위'])
ws.append([1,2,3,4,5])

ws.cell(2,1).fill=fill_color.ptn_red
ws.cell(2,2).fill=fill_color.ptn_blue
ws.cell(2,3).fill=fill_color.ptn_springgreen
ws.cell(2,4).fill=fill_color.ptn_yellowish
ws.cell(2,5).fill=fill_color.ptn_skyblue

font_20=Font(name='나눔고딕',size=20,color=font_color.skyblue2)
ws.cell(1,1).fill=fill_color.ptn_skyblue2
wb.save(r'C:\Users\xkr04\Downloads\excel\excel3.xlsx')

